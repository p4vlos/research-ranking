"""
Export results to Excel spreadsheet.
"""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)


def export_to_excel(df: pd.DataFrame, top10_data: dict, area_papers: dict):
    """Export ranking summary and top papers to Excel."""
    filepath = OUTPUT_DIR / "research_area_ranking.xlsx"

    # --- Sheet 1: Summary Ranking ---
    summary_cols = [
        "rank", "name", "paper_count", "papers_2024", "papers_2025",
        "growth_rate", "total_citations", "mean_citations",
        "median_citations", "max_citations", "top10_mean", "composite_score"
    ]
    summary_df = df[summary_cols].copy()
    summary_df.columns = [
        "Rank", "Research Area", "Total Papers", "Papers 2024", "Papers 2025",
        "Growth Rate (%)", "Total Citations", "Mean Citations",
        "Median Citations", "Max Citations", "Top-10 Mean Cit.", "Composite Score"
    ]

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        # Summary sheet
        summary_df.to_excel(writer, sheet_name="Ranking Summary", index=False)

        # Top-10 papers per area
        for _, row in df.iterrows():
            area_id = row["area_id"]
            area_short = row["short"][:20]
            papers = top10_data.get(area_id, [])

            if papers:
                top_df = pd.DataFrame(papers)
                top_df = top_df[["title", "citations", "year", "venue", "first_author"]]
                top_df.columns = ["Title", "Citations", "Year", "Venue", "First Author"]
                sheet_name = f"Top10_{area_short}"[:31]  # Excel 31 char limit
                top_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Raw counts per area (all papers)
        all_papers_list = []
        for _, row in df.iterrows():
            area_id = row["area_id"]
            for p in area_papers.get(area_id, []):
                all_papers_list.append({
                    "Research Area": row["name"],
                    "Title": p.get("title", ""),
                    "Year": p.get("year", ""),
                    "Citations": p.get("citations", 0),
                    "Venue": p.get("venue", ""),
                    "First Author": p.get("first_author", ""),
                    "DOI": p.get("doi", ""),
                })

        if all_papers_list:
            all_df = pd.DataFrame(all_papers_list)
            all_df = all_df.sort_values(
                ["Research Area", "Citations"], ascending=[True, False]
            )
            all_df.to_excel(writer, sheet_name="All Papers", index=False)

    # Format the workbook
    _format_workbook(filepath)

    print(f"\n  Saved: {filepath}")
    return filepath


def _format_workbook(filepath: Path):
    """Apply formatting to the Excel workbook."""
    wb = load_workbook(filepath)

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC")
    )

    for ws in wb.worksheets:
        # Format headers
        for col_idx, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    cell_len = len(str(cell.value or ""))
                    max_length = max(max_length, min(cell_len, 50))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 3

        # Add borders to data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

    # Highlight top-3 in summary sheet
    summary_ws = wb["Ranking Summary"]
    gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    bronze_fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")

    fills = [gold_fill, silver_fill, bronze_fill]
    for row_idx in range(2, min(5, summary_ws.max_row + 1)):
        fill = fills[row_idx - 2]
        for cell in summary_ws[row_idx]:
            cell.fill = fill
            cell.font = Font(bold=True, size=11)

    wb.save(filepath)
